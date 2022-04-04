from datetime import datetime

import openpyxl

from src.python import filepath, satnogs

TABS = {0: "allSatellite", 1: "filteredSatellite", 2: "sortedSatellite", 3: "TLE"}
FILE_DIR = filepath.getRoot() + "/CubeSAT/satnogs" + str(datetime.now().date()) + ".xlsx"
DATA_0 = satnogs.getSatellites()
DATA_1 = satnogs.satelliteFilter(DATA_0)
DATA_2 = satnogs.sortMostRecent(DATA_1)
TLE = satnogs.tleFilter(DATA_2)
DATA = [DATA_0, DATA_1, DATA_2, TLE]


def setContent(ws, data):
    for r in range(2, len(data) + 2):
        entry = data[r - 2]

        try:
            ws.cell(r, 1, entry["name"])
            ws.cell(r, 2, entry["description"])
            ws.cell(r, 3, entry["norad_cat_id"])
            ws.cell(r, 4, entry["service"])
            ws.cell(r, 5, entry["mode"])
            ws.cell(r, 6, entry["baud"])
            ws.cell(r, 7, entry["time"])
        except KeyError:
            ws["A1"] = "tle0"
            ws["B1"] = "tle1"
            ws["C1"] = "tle2"
            ws["D1"] = "tle_source"
            ws["E1"] = "norad_cat_id"
            ws["F1"] = "updated"

            for row in range(0, len(TLE)):
                for col in range(0, len(TLE[row])):
                    ws.cell(row + 2, col + 1, list(TLE[row].values())[col])


def setTab(wb: openpyxl.Workbook) -> None:
    for k in TABS:
        ws = wb.create_sheet(TABS[k], k)
        setHeader(ws)
        setContent(ws, DATA[k])


def setHeader(ws: openpyxl.Workbook.worksheets) -> None:
    ws["A1"] = "Name"
    ws["B1"] = "Description"
    ws["C1"] = "Norad_cat_id"
    ws["D1"] = "Service"
    ws["E1"] = "Mode"
    ws["F1"] = "Baud Rate"
    ws["G1"] = "Timestamp"


def setWorkbook():
    wb = openpyxl.Workbook()
    setTab(wb)
    wb.save(FILE_DIR)
    wb.close()


setWorkbook()
